"""
Generator for Luminary Systems Mock Meraki API seed data.

Reads assets_luminary.xlsx rows where seen_by includes 'meraki'
(SFO, NYC, LON sites) and produces JSON files consumed by app.py.

All IDs, serials, and MACs are deterministic (md5-derived) so a rebuild
never churns primary keys used in Postman / integrations.
"""

import hashlib
import json
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT     = Path(__file__).resolve().parent.parent
RAW      = ROOT / "seed_data" / "raw"
_CENTRAL = ROOT.parent / "luminary-demo-docs" / "master-sheet" / "assets_luminary.xlsx"
_LOCAL   = ROOT / "seed_data" / "source" / "assets_luminary.xlsx"
XLSX     = _CENTRAL if _CENTRAL.exists() else _LOCAL

NOW_EPOCH = int(time.time())  # current UTC epoch at generation time

# ---------------------------------------------------------------------------
# Deterministic helpers
# ---------------------------------------------------------------------------
def _md5(s: str) -> str:
    return hashlib.md5(s.encode()).hexdigest()

def make_uuid(seed: str) -> str:
    h = _md5(seed)
    return f"{h[0:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"

def make_org_id(seed: str) -> str:
    """Real Meraki org IDs are numeric strings (6 digits)."""
    return str(int(_md5(seed), 16) % 900000 + 100000)

def make_network_id(seed: str) -> str:
    """Real Meraki network IDs use N_ prefix + numeric suffix."""
    return "N_" + str(int(_md5(seed), 16) % 900000000 + 100000000)

def make_mac_colon(seed: str) -> str:
    """Meraki MAC format: aa:bb:cc:dd:ee:ff (lowercase colon-separated)."""
    h = _md5(seed)
    return ":".join(h[i:i+2] for i in range(0, 12, 2))

def make_serial(prefix: str, seed: str) -> str:
    """Meraki serial format: Q2XX-ABCD-EFGH (uppercase)."""
    h = _md5(seed).upper()
    return f"{prefix}-{h[0:4]}-{h[4:8]}"

def device_notes(name: str, site_cfg: dict) -> str:
    """Build the notes string from device name: lsys-{site}-f{N}-{type}-{num}."""
    import re
    m = re.search(r'-f(\d+)-', name)
    floor = int(m.group(1)) if m else None
    building = site_cfg.get("building", site_cfg["addr"])
    if floor:
        return f"{building}, Floor {floor}"
    return building

def make_ts_str(offset_secs: int) -> str:
    """Unix timestamp string (Meraki firstSeen/lastSeen format)."""
    return str(NOW_EPOCH - offset_secs)

def make_iso(offset_secs: int) -> str:
    """ISO 8601 string for device configurationUpdatedAt / lastReportedAt."""
    import datetime
    dt = datetime.datetime.utcfromtimestamp(NOW_EPOCH - offset_secs)
    return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

# ---------------------------------------------------------------------------
# Site & org definitions  (Meraki sites: SFO / NYC / LON)
# ---------------------------------------------------------------------------
ORG_ID = make_org_id("org:luminary-systems-meraki")

SITES = {
    "San Francisco": {
        "code":   "sfo",
        "octet":  11,
        "tz":     "America/Los_Angeles",
        "lat":    37.7749,
        "lng":    -122.4194,
        "addr":   "San Francisco",
        "building": "San Francisco HQ",
    },
    "New York": {
        "code":   "nyc",
        "octet":  12,
        "tz":     "America/New_York",
        "lat":    40.7128,
        "lng":    -74.0060,
        "addr":   "New York",
        "building": "New York Office",
    },
    "London": {
        "code":   "lon",
        "octet":  13,
        "tz":     "Europe/London",
        "lat":    51.5074,
        "lng":    -0.1278,
        "addr":   "London",
        "building": "London Office",
    },
}

for loc, cfg in SITES.items():
    cfg["network_id"] = make_network_id(f"network:meraki:{loc}")

# ---------------------------------------------------------------------------
# Category → Meraki product type / client type
# ---------------------------------------------------------------------------
DEVICE_CATS = {"wap", "switch", "router"}

# Wireless client categories (connected via AP on Corporate or IoT SSID)
WIRELESS_CATS = {"win_laptop", "mac_laptop", "mobile_ios", "mobile_droid", "iot"}

# Wired client categories (connected via switchport)
WIRED_CATS = {"linux_ws", "printer", "clinic"}

# Device models
DEVICE_MODEL = {
    "wap":    {"productType": "wireless",  "model": "MR57",      "serial_prefix": "Q2KD", "firmware_disp": "MR 29.6",  "firmware_api": "wireless-29-6"},
    "switch": {"productType": "switch",    "model": "MS225-48",  "serial_prefix": "Q2QN", "firmware_disp": "MS 15.21", "firmware_api": "switch-15-21"},
    "router": {"productType": "appliance", "model": "MX450",     "serial_prefix": "Q2QN", "firmware_disp": "MX 18.211","firmware_api": "wired-18-211"},
}

# Client os/manufacturer by category
CLIENT_CFG = {
    "win_laptop":   {"os": "Windows 11",      "manufacturer": "Dell",    "deviceTypePrediction": "Computer",    "ssid": "Corporate", "vlan": 1,  "namedVlan": "Corporate",  "recentDeviceConnection": "Wireless", "switchport": None},
    "mac_laptop":   {"os": "macOS Sonoma",    "manufacturer": "Apple",   "deviceTypePrediction": "Computer",    "ssid": "Corporate", "vlan": 1,  "namedVlan": "Corporate",  "recentDeviceConnection": "Wireless", "switchport": None},
    "mobile_ios":   {"os": "iOS 17",          "manufacturer": "Apple",   "deviceTypePrediction": "Phone",       "ssid": "Corporate", "vlan": 4,  "namedVlan": "Mobile",     "recentDeviceConnection": "Wireless", "switchport": None},
    "mobile_droid": {"os": "Android 14",      "manufacturer": "Samsung", "deviceTypePrediction": "Phone",       "ssid": "Corporate", "vlan": 4,  "namedVlan": "Mobile",     "recentDeviceConnection": "Wireless", "switchport": None},
    "iot":          {"os": "Linux",           "manufacturer": "Hikvision","deviceTypePrediction": "IP camera",  "ssid": "IoT",       "vlan": 8,  "namedVlan": "IoT",        "recentDeviceConnection": "Wireless", "switchport": None},
    "linux_ws":     {"os": "Linux",           "manufacturer": "Dell",    "deviceTypePrediction": "Computer",    "ssid": None,        "vlan": 1,  "namedVlan": "Corporate",  "recentDeviceConnection": "Wired",    "switchport": "1"},
    "printer":      {"os": None,              "manufacturer": "HP",      "deviceTypePrediction": "Printer",     "ssid": None,        "vlan": 5,  "namedVlan": "Printers",   "recentDeviceConnection": "Wired",    "switchport": "1"},
    "clinic":       {"os": None,              "manufacturer": "Philips", "deviceTypePrediction": "Medical",     "ssid": None,        "vlan": 9,  "namedVlan": "Clinic",     "recentDeviceConnection": "Wired",    "switchport": "1"},
}

# VLAN definitions per site (third octet = vlan segment)
VLAN_DEFS = [
    {"id": 1,  "segment": 1,  "name": "Corporate",   "desc": "User endpoints — laptops & workstations"},
    {"id": 4,  "segment": 4,  "name": "Mobile",      "desc": "Mobile devices — iOS & Android"},
    {"id": 5,  "segment": 5,  "name": "Printers",    "desc": "Office printers and MFPs"},
    {"id": 6,  "segment": 6,  "name": "Access Points","desc": "Wireless APs — management VLAN"},
    {"id": 8,  "segment": 8,  "name": "IoT",         "desc": "IoT and surveillance cameras"},
    {"id": 9,  "segment": 9,  "name": "Clinic",      "desc": "Clinical and medical devices"},
    {"id": 99, "segment": 0,  "name": "Management",  "desc": "Infrastructure management"},
]

# ---------------------------------------------------------------------------
# Load master sheet
# ---------------------------------------------------------------------------
print(f"Reading: {XLSX}")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
H = {h: i for i, h in enumerate(headers)}

rows_by_site = defaultdict(list)
for row in ws.iter_rows(min_row=2, values_only=True):
    loc  = row[H["location"]]
    sb   = row[H["seen_by"]] or ""
    cat  = row[H["category"]]
    if loc not in SITES:
        continue
    if "meraki" not in sb.lower():
        continue
    rows_by_site[loc].append({
        "hostname":    row[H["hostname"]],
        "ip":          row[H["ip_address"]],
        "mac":         row[H["mac_address"]],
        "category":    cat,
        "manufacturer":row[H["manufacturer"]],
        "model":       row[H["model"]],
        "serial":      row[H["serial"]],
        "assigned_to": row[H["assigned_to"]],
    })
wb.close()

total = sum(len(v) for v in rows_by_site.values())
print(f"Total Meraki rows: {total}")
for loc, rows in sorted(rows_by_site.items()):
    from collections import Counter
    cats = Counter(r["category"] for r in rows)
    print(f"  {loc}: {len(rows)} {dict(sorted(cats.items()))}")

# ---------------------------------------------------------------------------
# Build data structures
# ---------------------------------------------------------------------------
org = {
    "id":   ORG_ID,
    "name": "Luminary Systems",
    "url":  "https://mock.meraki.com/organizations/luminary-systems/manage",
    "api":  {"enabled": True},
    "licensing": {"model": "co-term"},
    "cloud": {
        "region": {
            "name": "North America",
            "host": {"name": "United States"}
        }
    },
    "management": {
        "details": [
            {"name": "customer number", "value": str(int(_md5("org:luminary:customer"), 16) % 90000000 + 10000000)}
        ]
    },
}

networks    = []
devices     = []           # org-level device list
clients_by_net = {}        # network_id → list of network clients
device_clients = {}        # serial → list of device-level clients
vlans_by_net   = {}        # network_id → list of VLANs

for loc, rows in rows_by_site.items():
    cfg = SITES[loc]
    net_id = cfg["network_id"]
    octet  = cfg["octet"]
    code   = cfg["code"]

    # -- Network --
    networks.append({
        "id":            net_id,
        "organizationId": ORG_ID,
        "name":          loc,
        "productTypes":  ["appliance", "switch", "wireless"],
        "timeZone":      cfg["tz"],
        "tags":          [code],
        "enrollmentString": None,
        "url":           f"https://mock.meraki.com/luminary-{code}/manage/clients",
        "notes":         None,
        "details":       None,
        "isBoundToConfigTemplate": False,
        "isVirtual":     False,
    })

    # -- VLANs --
    site_vlans = []
    for vd in VLAN_DEFS:
        seg     = vd["segment"]
        vlan_id = vd["id"]
        subnet  = f"10.{octet}.{seg}.0/24"
        app_ip  = f"10.{octet}.{seg}.1"
        iface_id = _md5(f"iface:{net_id}:{vlan_id}")[:13]
        site_vlans.append({
            "id":              str(vlan_id),
            "interfaceId":     iface_id,
            "networkId":       net_id,
            "name":            vd["name"],
            "applianceIp":     app_ip,
            "subnet":          subnet,
            "fixedIpAssignments": {},
            "reservedIpRanges":   [],
            "dnsNameservers":  "upstream_dns",
            "dhcpHandling":    "Run a DHCP server",
            "dhcpLeaseTime":   "1 day",
            "dhcpBootOptionsEnabled": False,
            "dhcpOptions":     [],
            "vpnNatSubnet":    subnet,
            "mandatoryDhcp":   {"enabled": False},
            "ipv6":            {"enabled": False},
            "dhcpBootFilename":    None,
            "dhcpBootNextServer":  None,
            "dhcpRelayServerIps":  [],
            "groupPolicyId":   None,
            "templateVlanType":"same",
            "cidr":            subnet,
            "mask":            24,
        })
    vlans_by_net[net_id] = site_vlans

    # -- Devices & Clients --
    net_clients  = []
    site_dev_rows = [r for r in rows if r["category"] in DEVICE_CATS]
    site_cli_rows = [r for r in rows if r["category"] not in DEVICE_CATS]

    # Track switch serials per site (for wired client recentDeviceSerial)
    switch_serials = []

    _TYPE_LABEL = {"wap": "ap", "switch": "sw", "router": "mx"}
    _dev_counters = {"wap": 0, "switch": 0, "router": 0}

    for r in site_dev_rows:
        cat     = r["category"]
        dm      = DEVICE_MODEL[cat]
        hostname = r["hostname"]
        seed    = f"meraki:device:{hostname}"  # keep seed for deterministic serial/MAC
        _dev_counters[cat] += 1
        meraki_name = hostname  # name comes from xlsx (already in lsys-{site}-f{N}-{type}-{##} format)
        serial  = make_serial(dm["serial_prefix"], seed)
        mac     = make_mac_colon(f"mac:device:{hostname}")
        # Use IP from xlsx or derive
        raw_ip  = str(r["ip"]) if r["ip"] else None
        lan_ip  = raw_ip if raw_ip and raw_ip.startswith("10.") else f"10.{octet}.6.{hash(hostname)%200+2}"

        offset_cfg = int(_md5(seed)[0:4], 16) % 86400
        device = {
            "serial":           serial,
            "name":             meraki_name,
            "mac":              mac,
            "networkId":        net_id,
            "organizationId":   ORG_ID,
            "model":            dm["model"],
            "productType":      dm["productType"],
            "firmware":         dm["firmware_api"],
            "lanIp":            lan_ip,
            "tags":             [code],
            "lat":              cfg["lat"] + int(_md5(seed)[4:8], 16) % 100 / 10000,
            "lng":              cfg["lng"] + int(_md5(seed)[8:12], 16) % 100 / 10000,
            "address":          cfg["addr"],
            "notes":            device_notes(meraki_name, cfg),
            "url":              f"https://mock.meraki.com/devices/{serial}/manage",
            "configurationUpdatedAt": make_iso(offset_cfg),
            "details":          [{"name": "Running software version", "value": dm["firmware_disp"]}],
            "_site":            loc,
            "_category":        cat,
        }
        if cat == "router":
            device["wan1Ip"] = f"203.{int(_md5(seed)[0:2],16)}.{int(_md5(seed)[2:4],16)}.{int(_md5(seed)[4:6],16)}"
        devices.append(device)
        if cat == "switch":
            switch_serials.append(serial)
        # Devices have no device-level clients of their own (APs and switches
        # forward clients; the client list is at the network level)
        device_clients[serial] = []

    # Now assign AP serials per site for wireless client's recentDeviceSerial
    ap_serials = [d["serial"] for d in devices if d.get("_site") == loc and d.get("_category") == "wap"]
    ap_names   = {d["serial"]: d["name"] for d in devices if d.get("_site") == loc and d.get("_category") == "wap"}
    ap_macs    = {d["serial"]: d["mac"]  for d in devices if d.get("_site") == loc and d.get("_category") == "wap"}

    sw_names   = {d["serial"]: d["name"] for d in devices if d.get("_site") == loc and d.get("_category") == "switch"}
    sw_macs    = {d["serial"]: d["mac"]  for d in devices if d.get("_site") == loc and d.get("_category") == "switch"}

    for idx, r in enumerate(site_cli_rows):
        cat     = r["category"]
        ccfg    = CLIENT_CFG.get(cat, CLIENT_CFG["win_laptop"])
        hostname = r["hostname"]
        seed    = f"meraki:client:{hostname}"
        h       = _md5(seed)

        # MAC: colon-separated lowercase (from xlsx or derived)
        raw_mac = str(r["mac"]).lower().replace("-","").replace(":","").replace(".","") if r["mac"] else None
        if raw_mac and len(raw_mac) == 12:
            mac = ":".join(raw_mac[i:i+2] for i in range(0,12,2))
        else:
            mac = make_mac_colon(seed)

        # IP from xlsx
        raw_ip = str(r["ip"]) if r["ip"] else None
        if raw_ip and raw_ip.startswith("10."):
            ip = raw_ip
        else:
            seg  = ccfg["vlan"] if ccfg["vlan"] != 1 else 1
            ip   = f"10.{octet}.{seg}.{(idx % 250) + 2}"

        # Timestamps
        offset_first = int(h[0:4], 16) % (86400 * 90)
        offset_last  = int(h[4:8], 16) % 3600
        first_seen_ts = make_ts_str(offset_first)
        last_seen_ts  = make_ts_str(offset_last)

        # Usage (bytes for network endpoint; KB for device endpoint)
        sent_bytes = int(h[8:12],  16) * 50000
        recv_bytes = int(h[12:16], 16) * 100000

        # Assign to AP or switch
        is_wired = ccfg["recentDeviceConnection"] == "Wired"
        if is_wired and switch_serials:
            dev_serial = switch_serials[int(h[16:18], 16) % len(switch_serials)]
            dev_name   = sw_names.get(dev_serial, "")
            dev_mac    = sw_macs.get(dev_serial, "")
            switchport = f"1/{int(h[18:20],16) % 48 + 1}"
        elif not is_wired and ap_serials:
            dev_serial = ap_serials[int(h[16:18], 16) % len(ap_serials)]
            dev_name   = ap_names.get(dev_serial, "")
            dev_mac    = ap_macs.get(dev_serial, "")
            switchport = None
        else:
            dev_serial = None
            dev_name   = None
            dev_mac    = None
            switchport = ccfg.get("switchport")

        client_id = h[:12]  # short id

        # User from assigned_to
        user = str(r["assigned_to"]) if r["assigned_to"] else None
        if user and "@" not in user:
            user = None

        net_client = {
            "id":                    client_id,
            "mac":                   mac,
            "ip":                    ip,
            "ip6":                   None,
            "ip6Local":              None,
            "description":           hostname,
            "firstSeen":             first_seen_ts,
            "lastSeen":              last_seen_ts,
            "manufacturer":          r["manufacturer"] or ccfg["manufacturer"],
            "os":                    ccfg["os"],
            "deviceTypePrediction":  ccfg["deviceTypePrediction"],
            "user":                  user,
            "vlan":                  str(ccfg["vlan"]),
            "namedVlan":             ccfg["namedVlan"],
            "ssid":                  ccfg["ssid"],
            "switchport":            switchport,
            "wirelessCapabilities":  "802.11ac - 2.4 and 5 GHz" if not is_wired else None,
            "smInstalled":           False,
            "recentDeviceSerial":    dev_serial,
            "recentDeviceName":      dev_name,
            "recentDeviceMac":       dev_mac,
            "recentDeviceConnection":ccfg["recentDeviceConnection"],
            "notes":                 None,
            "groupPolicy8021x":      None,
            "adaptivePolicyGroup":   None,
            "pskGroup":              None,
            "status":                "Online",
            "usage":                 {"sent": sent_bytes, "recv": recv_bytes, "total": sent_bytes + recv_bytes},
            "_network_id":           net_id,
        }
        net_clients.append(net_client)

        # Also add to the device-level client list (simpler schema)
        if dev_serial and dev_serial in device_clients:
            device_clients[dev_serial].append({
                "id":                client_id,
                "mac":               mac,
                "description":       hostname,
                "mdnsName":          hostname,
                "dhcpHostname":      hostname.upper()[:15],
                "user":              user,
                "ip":                ip,
                "vlan":              str(ccfg["vlan"]),
                "namedVlan":         ccfg["namedVlan"],
                "switchport":        switchport,
                "adaptivePolicyGroup": None,
                "usage": {
                    "sent": sent_bytes // 1000,
                    "recv": recv_bytes // 1000,
                },
            })

    clients_by_net[net_id] = net_clients

# ---------------------------------------------------------------------------
# Build device_statuses and device_availabilities
# ---------------------------------------------------------------------------
device_statuses       = []
device_availabilities = []

for d in devices:
    seed   = f"status:{d['serial']}"
    h      = _md5(seed)
    status = "online"  # All devices online (stable demo)

    lan_ip = d.get("lanIp", "10.0.0.1")
    gw_parts = lan_ip.rsplit(".", 1)
    gateway  = f"{gw_parts[0]}.1"

    offset_reported = int(h[0:4], 16) % 300  # within last 5 min
    last_reported   = make_iso(offset_reported)
    # Microseconds
    last_reported   = last_reported.replace("Z", f".{int(h[4:8],16) % 1000000:06d}Z")

    device_statuses.append({
        "name":           d["name"],
        "serial":         d["serial"],
        "mac":            d["mac"],
        "publicIp":       d.get("wan1Ip", f"203.{int(h[0:2],16)}.{int(h[2:4],16)}.{int(h[4:6],16)}"),
        "networkId":      d["networkId"],
        "status":         status,
        "lastReportedAt": last_reported,
        "lanIp":          lan_ip,
        "gateway":        gateway,
        "ipType":         "static",
        "primaryDns":     "8.8.8.8",
        "secondaryDns":   "8.8.4.4",
        "productType":    d["productType"],
        "model":          d["model"],
        "tags":           d.get("tags", []),
        "components":     {"powerSupplies": []},
    })

    device_availabilities.append({
        "serial":      d["serial"],
        "name":        d["name"],
        "mac":         d["mac"],
        "network":     {"id": d["networkId"]},
        "productType": d["productType"],
        "status":      status,
        "tags":        d.get("tags", []),
    })

# ---------------------------------------------------------------------------
# Strip internal fields before writing
# ---------------------------------------------------------------------------
devices_clean = []
for d in devices:
    dc = {k: v for k, v in d.items() if not k.startswith("_")}
    devices_clean.append(dc)

clients_clean = {}
for net_id, clients in clients_by_net.items():
    clients_clean[net_id] = [
        {k: v for k, v in c.items() if not k.startswith("_")}
        for c in clients
    ]

# ---------------------------------------------------------------------------
# Print summary
# ---------------------------------------------------------------------------
from collections import Counter
print(f"\nOrg: {ORG_ID}")
print(f"Networks: {len(networks)}")
print(f"Devices: {len(devices_clean)}")
prod_counts = Counter(d["productType"] for d in devices_clean)
for pt, cnt in sorted(prod_counts.items()):
    print(f"  {pt}: {cnt}")
total_clients = sum(len(v) for v in clients_clean.values())
print(f"Clients (network-level): {total_clients}")
for net in networks:
    nc = len(clients_clean.get(net["id"], []))
    nd = sum(1 for d in devices_clean if d["networkId"] == net["id"])
    print(f"  {net['name']}: {nc} clients + {nd} devices")

# ---------------------------------------------------------------------------
# Write JSON files
# ---------------------------------------------------------------------------
RAW.mkdir(parents=True, exist_ok=True)

(RAW / "org.json").write_text(json.dumps(org, indent=2))
(RAW / "networks.json").write_text(json.dumps(networks, indent=2))
(RAW / "devices.json").write_text(json.dumps(devices_clean, indent=2))
(RAW / "device_statuses.json").write_text(json.dumps(device_statuses, indent=2))
(RAW / "device_availabilities.json").write_text(json.dumps(device_availabilities, indent=2))
(RAW / "clients.json").write_text(json.dumps(clients_clean, indent=2))
(RAW / "device_clients.json").write_text(json.dumps(device_clients, indent=2))
(RAW / "vlans.json").write_text(json.dumps(vlans_by_net, indent=2))

print(f"\nWrote to {RAW}/")
print("  org.json, networks.json, devices.json")
print("  device_statuses.json, device_availabilities.json")
print("  clients.json, device_clients.json, vlans.json")
